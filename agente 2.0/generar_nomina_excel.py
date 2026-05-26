from __future__ import annotations

from datetime import date, timedelta
import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


OUT = Path("Nomina_Agentica_Colombia_2026.xlsx")
random.seed(20260526)


PARAMS = {
    "Anio": 2026,
    "Mes": 5,
    "DiasPeriodo": 30,
    "SMLMV": 1750905,
    "AuxTransporte": 249095,
    "TopeAuxTransporte": 3501810,
    "HorasMes": 240,
    "SaludEmpleado": 0.04,
    "PensionEmpleado": 0.04,
    "RiesgoValidacionSalarioAlto": 20000000,
    "FactorHoraExtraDiurna": 1.25,
    "FactorHoraExtraNocturna": 1.75,
    "FactorRecargoNocturno": 0.35,
    "FactorDominicalFestivo": 1.75,
    "FuenteSMLMV": "Decreto 1469 de 2025 / Presidencia Colombia",
    "FuenteAuxTransporte": "Decreto 1470 de 2025 / SUIN Juriscol",
}

FIRST_NAMES = [
    "Valentina", "Santiago", "Camila", "Mateo", "Isabella", "Daniel", "Mariana", "Sebastian",
    "Laura", "Nicolas", "Sofia", "Juan", "Ana", "Andres", "Paula", "David", "Carolina",
    "Felipe", "Juliana", "Miguel", "Catalina", "Alejandro", "Manuela", "Diego", "Luisa",
    "Tomas", "Daniela", "Carlos", "Natalia", "Jorge",
]
LAST_NAMES = [
    "Gomez", "Rodriguez", "Martinez", "Garcia", "Lopez", "Hernandez", "Diaz", "Perez",
    "Sanchez", "Ramirez", "Torres", "Vargas", "Castro", "Rojas", "Moreno", "Jimenez",
]
CARGOS = [
    ("Auxiliar Operativo", 1750905),
    ("Asistente Administrativo", 2100000),
    ("Analista de Nomina", 3200000),
    ("Coordinador Comercial", 4500000),
    ("Desarrollador Junior", 3800000),
    ("Desarrollador Senior", 8500000),
    ("Contador", 5200000),
    ("Jefe de Talento Humano", 7800000),
    ("Vendedor", 2300000),
    ("Supervisor de Planta", 3600000),
]
EPS = ["Sura", "Sanitas", "Compensar", "Nueva EPS", "Famisanar", "Salud Total"]
PENSION = ["Porvenir", "Proteccion", "Colfondos", "Skandia", "Colpensiones"]
CONTRATOS = ["Indefinido", "Fijo", "Obra labor", "Aprendizaje"]
AREAS = ["Operaciones", "Administracion", "Tecnologia", "Comercial", "Finanzas", "Talento Humano"]


def money(value: int | float) -> int:
    return int(round(value, 0))


def style_sheet(ws, freeze="A2"):
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def add_table(ws, name, ref):
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)


def set_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def header(ws, row=1):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def make_workbook():
    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    money_fmt = '"$"#,##0'
    pct_fmt = "0.00%"

    ws_p = wb.create_sheet("Parametros")
    ws_p.append(["Parametro", "Valor", "Descripcion"])
    rows = [
        ("Anio", PARAMS["Anio"], "Ano fiscal editable"),
        ("Mes", PARAMS["Mes"], "Mes de liquidacion 1-12"),
        ("DiasPeriodo", PARAMS["DiasPeriodo"], "Dias liquidados por periodo"),
        ("SMLMV", PARAMS["SMLMV"], "Salario minimo legal mensual vigente 2026"),
        ("AuxTransporte", PARAMS["AuxTransporte"], "Auxilio legal de transporte mensual 2026"),
        ("TopeAuxTransporte", PARAMS["TopeAuxTransporte"], "Aplica si salario base <= 2 SMLMV"),
        ("HorasMes", PARAMS["HorasMes"], "Base de horas mensuales para valor hora"),
        ("SaludEmpleado", PARAMS["SaludEmpleado"], "Aporte salud empleado"),
        ("PensionEmpleado", PARAMS["PensionEmpleado"], "Aporte pension empleado"),
        ("RiesgoValidacionSalarioAlto", PARAMS["RiesgoValidacionSalarioAlto"], "Umbral de alerta por salario alto"),
        ("FactorHoraExtraDiurna", PARAMS["FactorHoraExtraDiurna"], "Factor sobre hora ordinaria"),
        ("FactorHoraExtraNocturna", PARAMS["FactorHoraExtraNocturna"], "Factor sobre hora ordinaria"),
        ("FactorRecargoNocturno", PARAMS["FactorRecargoNocturno"], "Solo recargo adicional"),
        ("FactorDominicalFestivo", PARAMS["FactorDominicalFestivo"], "Factor sobre hora ordinaria"),
        ("FuenteSMLMV", PARAMS["FuenteSMLMV"], "Referencia normativa cargada en mayo de 2026"),
        ("FuenteAuxTransporte", PARAMS["FuenteAuxTransporte"], "Referencia normativa cargada en mayo de 2026"),
    ]
    for r in rows:
        ws_p.append(r)
    header(ws_p)
    add_table(ws_p, "tblParametros", f"A1:C{ws_p.max_row}")
    set_widths(ws_p, {"A": 28, "B": 24, "C": 56})
    for row in ws_p.iter_rows(min_row=2, max_col=3):
        for cell in row:
            cell.border = border
        if isinstance(row[1].value, float) and row[1].value < 1:
            row[1].number_format = pct_fmt
        elif isinstance(row[1].value, int):
            row[1].number_format = money_fmt if row[0].value not in {"Anio", "Mes", "DiasPeriodo", "HorasMes"} else "0"
    style_sheet(ws_p)

    ws_e = wb.create_sheet("Empleados")
    emp_headers = [
        "ID", "Nombre", "Cargo", "Area", "SalarioBase", "TipoContrato", "HorasContratadas",
        "FechaIngreso", "EPS", "Pension", "Ciudad", "Activo", "ObservacionAgente",
    ]
    ws_e.append(emp_headers)
    empleados = []
    used_names = set()
    for i in range(1, 31):
        cargo, base = random.choice(CARGOS)
        name = f"{random.choice(FIRST_NAMES)} {random.choice(LAST_NAMES)}"
        while name in used_names:
            name = f"{random.choice(FIRST_NAMES)} {random.choice(LAST_NAMES)}"
        used_names.add(name)
        salario = money(base * random.choice([1, 1, 1.05, 1.1, 1.2, 0.95]))
        if i <= 5:
            salario = PARAMS["SMLMV"]
        ingreso = date(2018, 1, 1) + timedelta(days=random.randint(0, 2950))
        empleado = [
            f"EMP-{i:04d}", name, cargo, random.choice(AREAS), salario, random.choice(CONTRATOS),
            240, ingreso, random.choice(EPS), random.choice(PENSION), random.choice(["Bogota", "Medellin", "Cali", "Barranquilla"]),
            "Si", "",
        ]
        empleados.append(empleado)
        ws_e.append(empleado)
    header(ws_e)
    add_table(ws_e, "tblEmpleados", f"A1:M{ws_e.max_row}")
    set_widths(ws_e, {"A": 12, "B": 24, "C": 26, "D": 18, "E": 15, "F": 16, "G": 16, "H": 14, "I": 16, "J": 16, "K": 14, "L": 10, "M": 36})
    for row in ws_e.iter_rows(min_row=2):
        row[4].number_format = money_fmt
        row[7].number_format = "yyyy-mm-dd"
        for cell in row:
            cell.border = border
    for col, opts in [("F", CONTRATOS), ("I", EPS), ("J", PENSION), ("L", ["Si", "No"])]:
        dv = DataValidation(type="list", formula1=f'"{",".join(opts)}"', allow_blank=False)
        ws_e.add_data_validation(dv)
        dv.add(f"{col}2:{col}500")
    ws_e.conditional_formatting.add("E2:E500", CellIsRule(operator="lessThan", formula=["Parametros!$B$5"], fill=PatternFill("solid", fgColor="FFC7CE")))
    ws_e.conditional_formatting.add("M2:M500", FormulaRule(formula=['LEN($M2)>0'], fill=PatternFill("solid", fgColor="FFF2CC")))
    style_sheet(ws_e)

    ws_n = wb.create_sheet("Novedades_Horas")
    nov_headers = [
        "ID", "EmpleadoID", "HorasOrdinarias", "HorasExtraDiurnas", "HorasExtraNocturnas",
        "HorasRecargoNocturno", "HorasDominicalFestivo", "DiasNoTrabajados", "Bonificaciones",
        "PrestamosDescuentos", "Observacion",
    ]
    ws_n.append(nov_headers)
    for idx, emp in enumerate(empleados, start=1):
        ws_n.append([
            f"NOV-{idx:04d}", emp[0], 240 - random.choice([0, 0, 0, 8, 16]),
            random.choice([0, 0, 2, 4, 6, 8]),
            random.choice([0, 0, 1, 2, 4]),
            random.choice([0, 0, 5, 8, 10]),
            random.choice([0, 0, 0, 8]),
            random.choice([0, 0, 0, 1, 2]),
            money(random.choice([0, 0, 0, 150000, 300000, 500000])),
            money(random.choice([0, 0, 0, 100000, 250000])),
            "",
        ])
    header(ws_n)
    add_table(ws_n, "tblNovedades", f"A1:K{ws_n.max_row}")
    set_widths(ws_n, {"A": 12, "B": 13, "C": 16, "D": 18, "E": 20, "F": 22, "G": 22, "H": 17, "I": 15, "J": 19, "K": 34})
    for row in ws_n.iter_rows(min_row=2):
        for c in row:
            c.border = border
        row[8].number_format = money_fmt
        row[9].number_format = money_fmt
    style_sheet(ws_n)

    ws_c = wb.create_sheet("Calculo_Nomina")
    calc_headers = [
        "EmpleadoID", "Nombre", "Cargo", "Area", "SalarioBase", "ValorHora", "SalarioDevengado",
        "AuxTransporte", "ValorExtrasRecargos", "Bonificaciones", "TotalDevengado",
        "Salud", "Pension", "RetencionFuente", "PrestamosDescuentos", "TotalDeducciones",
        "NetoAPagar", "CostoEmpresaEstimado", "EstadoAgente",
    ]
    ws_c.append(calc_headers)
    for r in range(2, 32):
        emp_row = r
        nov_row = r
        ws_c.append([
            f"=Empleados!A{emp_row}",
            f"=Empleados!B{emp_row}",
            f"=Empleados!C{emp_row}",
            f"=Empleados!D{emp_row}",
            f"=Empleados!E{emp_row}",
            f"=E{r}/Parametros!$B$8",
            f"=ROUND(E{r}/Parametros!$B$8*Novedades_Horas!C{nov_row},0)",
            f'=IF(AND(E{r}<=Parametros!$B$7,Empleados!L{emp_row}="Si"),Parametros!$B$6,0)',
            (
                f"=ROUND(F{r}*(Novedades_Horas!D{nov_row}*Parametros!$B$12+"
                f"Novedades_Horas!E{nov_row}*Parametros!$B$13+"
                f"Novedades_Horas!F{nov_row}*Parametros!$B$14+"
                f"Novedades_Horas!G{nov_row}*Parametros!$B$15),0)"
            ),
            f"=Novedades_Horas!I{nov_row}",
            f"=SUM(G{r}:J{r})",
            f"=ROUND((G{r}+I{r}+J{r})*Parametros!$B$9,0)",
            f"=ROUND((G{r}+I{r}+J{r})*Parametros!$B$10,0)",
            f"=IF(K{r}>6000000,ROUND((K{r}-6000000)*0.1,0),0)",
            f"=Novedades_Horas!J{nov_row}",
            f"=SUM(L{r}:O{r})",
            f"=K{r}-P{r}",
            f"=ROUND(K{r}+(G{r}+I{r}+J{r})*0.205,0)",
            f'=IF(Q{r}<0,"ERROR: Neto negativo",IF(E{r}<Parametros!$B$5,"ERROR: Salario bajo",IF(E{r}>Parametros!$B$11,"REVISAR: salario alto","OK")))',
        ])
    header(ws_c)
    add_table(ws_c, "tblCalculoNomina", f"A1:S{ws_c.max_row}")
    set_widths(ws_c, {get_column_letter(i): 16 for i in range(1, 20)})
    set_widths(ws_c, {"B": 24, "C": 26, "S": 24})
    for row in ws_c.iter_rows(min_row=2):
        for c in row:
            c.border = border
        for idx in range(5, 19):
            row[idx - 1].number_format = money_fmt
    ws_c.conditional_formatting.add("S2:S500", FormulaRule(formula=['LEFT($S2,5)="ERROR"'], fill=PatternFill("solid", fgColor="FFC7CE")))
    ws_c.conditional_formatting.add("S2:S500", FormulaRule(formula=['LEFT($S2,7)="REVISAR"'], fill=PatternFill("solid", fgColor="FFF2CC")))
    style_sheet(ws_c)

    ws_v = wb.create_sheet("Validaciones")
    ws_v.append(["Regla", "Resultado", "Accion sugerida"])
    checks = [
        ("Empleados sin nombre", '=COUNTBLANK(Empleados!B2:B500)', "Completar nombre o eliminar registro vacio"),
        ("Salarios por debajo de SMLMV", '=COUNTIF(Empleados!E2:E500,"<"&Parametros!$B$5)', "Ajustar salario base a minimo legal o justificar excepcion"),
        ("Horas ordinarias fuera de rango", '=COUNTIFS(Novedades_Horas!C2:C500,">240")+COUNTIFS(Novedades_Horas!C2:C500,"<0")', "Revisar horas ordinarias del periodo"),
        ("Netos negativos", '=COUNTIF(Calculo_Nomina!Q2:Q500,"<0")', "Revisar deducciones, prestamos o novedades"),
        ("Registros marcados por agente", '=COUNTIF(Calculo_Nomina!S2:S500,"<>OK")', "Filtrar EstadoAgente en Calculo_Nomina"),
    ]
    for row in checks:
        ws_v.append(row)
    header(ws_v)
    add_table(ws_v, "tblValidaciones", f"A1:C{ws_v.max_row}")
    set_widths(ws_v, {"A": 34, "B": 16, "C": 58})
    ws_v.conditional_formatting.add("B2:B100", CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="FFC7CE")))
    style_sheet(ws_v)

    ws_r = wb.create_sheet("Reporte_Gerencial")
    ws_r["A1"] = "Resumen de nomina"
    ws_r["A1"].font = Font(size=18, bold=True, color="1F4E78")
    report_rows = [
        ("Periodo", '=DATE(Parametros!$B$2,Parametros!$B$3,1)'),
        ("Empleados activos", '=COUNTIF(Empleados!L:L,"Si")'),
        ("Total devengado", '=SUM(Calculo_Nomina!K:K)'),
        ("Total deducciones", '=SUM(Calculo_Nomina!P:P)'),
        ("Total neto a pagar", '=SUM(Calculo_Nomina!Q:Q)'),
        ("Costo empresa estimado", '=SUM(Calculo_Nomina!R:R)'),
        ("Alertas abiertas", '=COUNTIF(Calculo_Nomina!S:S,"<>OK")-1'),
    ]
    for idx, row in enumerate(report_rows, start=3):
        ws_r.cell(idx, 1, row[0])
        ws_r.cell(idx, 2, row[1])
        ws_r.cell(idx, 1).font = Font(bold=True)
    for row in range(3, 10):
        ws_r.cell(row, 2).number_format = money_fmt if row >= 5 else "0"
    ws_r["B3"].number_format = "mmmm yyyy"
    ws_r["D2"] = "Costo por area"
    ws_r["D2"].font = Font(bold=True, color="1F4E78")
    areas = sorted(set(AREAS))
    ws_r.append([])
    ws_r["D3"] = "Area"
    ws_r["E3"] = "Neto a pagar"
    ws_r["F3"] = "Costo estimado"
    for idx, area in enumerate(areas, start=4):
        ws_r.cell(idx, 4, area)
        ws_r.cell(idx, 5, f'=SUMIF(Calculo_Nomina!D:D,D{idx},Calculo_Nomina!Q:Q)')
        ws_r.cell(idx, 6, f'=SUMIF(Calculo_Nomina!D:D,D{idx},Calculo_Nomina!R:R)')
        ws_r.cell(idx, 5).number_format = money_fmt
        ws_r.cell(idx, 6).number_format = money_fmt
    header(ws_r, 3)
    set_widths(ws_r, {"A": 26, "B": 20, "D": 22, "E": 18, "F": 18})
    bar = BarChart()
    bar.title = "Neto a pagar por area"
    bar.y_axis.title = "COP"
    bar.x_axis.title = "Area"
    data = Reference(ws_r, min_col=5, min_row=3, max_row=3 + len(areas))
    cats = Reference(ws_r, min_col=4, min_row=4, max_row=3 + len(areas))
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.height = 8
    bar.width = 16
    ws_r.add_chart(bar, "H3")
    pie = PieChart()
    pie.title = "Distribucion costo estimado"
    pdata = Reference(ws_r, min_col=6, min_row=3, max_row=3 + len(areas))
    pie.add_data(pdata, titles_from_data=True)
    pie.set_categories(cats)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.height = 8
    pie.width = 12
    ws_r.add_chart(pie, "H20")
    style_sheet(ws_r, "A3")

    ws_s = wb.create_sheet("Escenarios")
    ws_s.append(["Escenario", "Parametro", "Valor", "Resultado estimado", "Formula/accion"])
    scenario_rows = [
        ("Base actual", "Aumento salarial", 0, '=SUM(Calculo_Nomina!Q:Q)', "Neto a pagar actual"),
        ("Simular aumento 10%", "Aumento salarial", 0.10, '=ROUND(SUM(Calculo_Nomina!Q:Q)*(1+C3),0)', "Impacto aproximado sobre neto"),
        ("Agregar 5 empleados minimo", "Cantidad", 5, '=C4*Parametros!$B$5', "Costo salarial base antes de aportes"),
        ("Sin auxilio transporte", "Auxilio", 0, '=SUM(Calculo_Nomina!Q:Q)-SUM(Calculo_Nomina!H:H)', "Comparativo sin auxilio"),
    ]
    for row in scenario_rows:
        ws_s.append(row)
    header(ws_s)
    add_table(ws_s, "tblEscenarios", f"A1:E{ws_s.max_row}")
    set_widths(ws_s, {"A": 28, "B": 24, "C": 16, "D": 22, "E": 42})
    for row in ws_s.iter_rows(min_row=2):
        row[2].number_format = "0.00%" if row[1].value == "Aumento salarial" else "0"
        row[3].number_format = money_fmt
    style_sheet(ws_s)

    ws_a = wb.create_sheet("Agente_NL")
    ws_a["A1"] = "Interfaz inteligente: comandos soportados"
    ws_a["A1"].font = Font(size=16, bold=True, color="1F4E78")
    ws_a.append(["Comando natural", "Como se ejecuta en el libro", "Estado"])
    commands = [
        ("Agrega 5 empleados con salario minimo", "Duplicar 5 filas base en Empleados con SalarioBase = Parametros!SMLMV y crear sus novedades", "Plantilla lista"),
        ("Calcula nomina del mes", "Editar Parametros Mes/Anio; Excel recalcula Calculo_Nomina y Reporte_Gerencial", "Automatico"),
        ("Simula aumento del 10%", "Usar fila Simular aumento 10% en Escenarios o cambiar C3", "Automatico"),
        ("Revisa errores", "Abrir Validaciones y filtrar EstadoAgente en Calculo_Nomina", "Automatico"),
        ("Genera reporte de fin de mes", "Reporte_Gerencial consolida totales, costos por area y graficos", "Automatico"),
    ]
    for row in commands:
        ws_a.append(row)
    header(ws_a, 2)
    set_widths(ws_a, {"A": 34, "B": 72, "C": 18})
    ws_a["A10"] = "Nota operativa"
    ws_a["A10"].font = Font(bold=True, color="1F4E78")
    ws_a["A11"] = "Este archivo usa formulas, tablas, validaciones y recalculo al abrir. Para eventos 100% activos tipo macro, se puede migrar a .xlsm u Office Scripts sobre esta misma estructura."
    ws_a["A11"].alignment = Alignment(wrap_text=True)
    ws_a.merge_cells("A11:C12")
    style_sheet(ws_a, "A3")

    ws_b = wb.create_sheet("Bitacora_Workflows")
    ws_b.append(["Workflow", "Disparador", "Accion automatizada", "Implementacion actual"])
    flows = [
        ("Alta empleado", "Nueva fila en Empleados", "Generar registro completo y novedades", "Tabla estructurada + validaciones + plantilla en Agente_NL"),
        ("Cambio salario", "Cambio en SalarioBase", "Recalcular devengado, deducciones, neto y alertas", "Formulas enlazadas con recalculo automatico"),
        ("Fin de mes", "Cambio Mes/Anio o cierre periodo", "Actualizar resumen y graficos", "Reporte_Gerencial con formulas y graficos"),
        ("Control errores", "Datos fuera de rango", "Marcar errores y sugerir correccion", "Validaciones + EstadoAgente + formato condicional"),
        ("Simulacion", "Cambio de escenario", "Estimar impacto financiero", "Hoja Escenarios"),
    ]
    for row in flows:
        ws_b.append(row)
    header(ws_b)
    add_table(ws_b, "tblWorkflows", f"A1:D{ws_b.max_row}")
    set_widths(ws_b, {"A": 22, "B": 30, "C": 44, "D": 46})
    style_sheet(ws_b)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
        ws.sheet_properties.pageSetUpPr.fitToPage = True

    wb.save(OUT)


if __name__ == "__main__":
    make_workbook()
    print(OUT.resolve())
