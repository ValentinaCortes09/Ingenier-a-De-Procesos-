from __future__ import annotations

import json
import os
import random
import re
import shutil
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import date
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
EXCEL_PATH = BASE_DIR / "Nomina_Agentica_Colombia_2026.xlsx"
BACKUP_DIR = BASE_DIR / "backups"
TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN", "").strip()
ALLOWED_CHAT_ID = os.environ.get("TELEGRAM_ALLOWED_CHAT_ID", "").strip()

MIN_WAGE = 1_750_905
TRANSPORT_ALLOWANCE = 249_095
TRANSPORT_LIMIT = MIN_WAGE * 2
HOURS_MONTH = 240

FIRST_NAMES = [
    "Valentina", "Santiago", "Camila", "Mateo", "Isabella", "Daniel", "Mariana", "Sebastian",
    "Laura", "Nicolas", "Sofia", "Juan", "Ana", "Andres", "Paula", "David", "Carolina",
    "Felipe", "Juliana", "Miguel", "Catalina", "Alejandro", "Manuela", "Diego", "Luisa",
]
LAST_NAMES = [
    "Gomez", "Rodriguez", "Martinez", "Garcia", "Lopez", "Hernandez", "Diaz", "Perez",
    "Sanchez", "Ramirez", "Torres", "Vargas", "Castro", "Rojas", "Moreno", "Jimenez",
]
EPS = ["Sura", "Sanitas", "Compensar", "Nueva EPS", "Famisanar", "Salud Total"]
PENSION = ["Porvenir", "Proteccion", "Colfondos", "Skandia", "Colpensiones"]
AREAS = ["Operaciones", "Administracion", "Tecnologia", "Comercial", "Finanzas", "Talento Humano"]


def api_url(method: str) -> str:
    return f"https://api.telegram.org/bot{TOKEN}/{method}"


def request_json(method: str, payload: dict | None = None) -> dict:
    data = None
    headers = {"Content-Type": "application/json"}
    if payload is not None:
        data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(api_url(method), data=data, headers=headers)
    with urllib.request.urlopen(req, timeout=60) as response:
        return json.loads(response.read().decode("utf-8"))


def send_message(chat_id: int, text: str) -> None:
    request_json("sendMessage", {"chat_id": chat_id, "text": text[:3900]})


def send_document(chat_id: int, path: Path, caption: str = "") -> None:
    boundary = f"----codex-{uuid4().hex}"
    fields = {
        "chat_id": str(chat_id),
        "caption": caption,
    }
    body = bytearray()
    for name, value in fields.items():
        body.extend(f"--{boundary}\r\n".encode())
        body.extend(f'Content-Disposition: form-data; name="{name}"\r\n\r\n'.encode())
        body.extend(str(value).encode("utf-8"))
        body.extend(b"\r\n")
    body.extend(f"--{boundary}\r\n".encode())
    body.extend(
        f'Content-Disposition: form-data; name="document"; filename="{path.name}"\r\n'
        "Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n"
        .encode()
    )
    body.extend(path.read_bytes())
    body.extend(b"\r\n")
    body.extend(f"--{boundary}--\r\n".encode())
    req = urllib.request.Request(
        api_url("sendDocument"),
        data=bytes(body),
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
    )
    with urllib.request.urlopen(req, timeout=120) as response:
        json.loads(response.read().decode("utf-8"))


def authorize(chat_id: int) -> bool:
    return not ALLOWED_CHAT_ID or str(chat_id) == ALLOWED_CHAT_ID


def backup_excel() -> Path:
    BACKUP_DIR.mkdir(exist_ok=True)
    stamp = time.strftime("%Y%m%d_%H%M%S")
    target = BACKUP_DIR / f"{EXCEL_PATH.stem}_{stamp}.xlsx"
    shutil.copy2(EXCEL_PATH, target)
    return target


def load_book():
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"No existe {EXCEL_PATH}")
    return load_workbook(EXCEL_PATH)


def save_book(wb) -> None:
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(EXCEL_PATH)


def next_employee_id(ws) -> str:
    ids = []
    for value in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        if value[0] and re.match(r"EMP-\d+", str(value[0])):
            ids.append(int(str(value[0]).split("-")[1]))
    return f"EMP-{(max(ids) if ids else 0) + 1:04d}"


def next_novelty_id(ws) -> str:
    ids = []
    for value in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        if value[0] and re.match(r"NOV-\d+", str(value[0])):
            ids.append(int(str(value[0]).split("-")[1]))
    return f"NOV-{(max(ids) if ids else 0) + 1:04d}"


def random_name() -> str:
    return f"{random.choice(FIRST_NAMES)} {random.choice(LAST_NAMES)}"


def add_minimum_wage_employees(count: int) -> str:
    count = max(1, min(count, 50))
    backup_excel()
    wb = load_book()
    ws_e = wb["Empleados"]
    ws_n = wb["Novedades_Horas"]
    created = []
    for _ in range(count):
        emp_id = next_employee_id(ws_e)
        ws_e.append([
            emp_id,
            random_name(),
            "Auxiliar Operativo",
            random.choice(AREAS),
            MIN_WAGE,
            "Indefinido",
            HOURS_MONTH,
            date.today(),
            random.choice(EPS),
            random.choice(PENSION),
            "Bogota",
            "Si",
            "Creado desde Telegram",
        ])
        ws_n.append([
            next_novelty_id(ws_n),
            emp_id,
            HOURS_MONTH,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            "Creado desde Telegram",
        ])
        created.append(emp_id)
    save_book(wb)
    return f"Listo. Agregue {count} empleado(s) con salario minimo: {', '.join(created)}."


def change_salary(emp_id: str, salary: int) -> str:
    backup_excel()
    wb = load_book()
    ws = wb["Empleados"]
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row, 1).value).upper() == emp_id.upper():
            old = ws.cell(row, 5).value
            ws.cell(row, 5).value = salary
            ws.cell(row, 13).value = f"Salario actualizado desde Telegram: {old} -> {salary}"
            save_book(wb)
            return f"Listo. Cambie el salario de {emp_id} de ${old:,.0f} a ${salary:,.0f}."
    return f"No encontre el empleado {emp_id}."


def calculate_row(salary: float, hours: float, hed: float, hen: float, hrn: float, hdf: float, bonus: float, loan: float) -> dict:
    value_hour = salary / HOURS_MONTH
    earned_salary = round(value_hour * hours)
    transport = TRANSPORT_ALLOWANCE if salary <= TRANSPORT_LIMIT else 0
    extras = round(value_hour * (hed * 1.25 + hen * 1.75 + hrn * 0.35 + hdf * 1.75))
    accrued = earned_salary + transport + extras + bonus
    base_social = earned_salary + extras + bonus
    health = round(base_social * 0.04)
    pension = round(base_social * 0.04)
    withholding = round((accrued - 6_000_000) * 0.1) if accrued > 6_000_000 else 0
    deductions = health + pension + withholding + loan
    net = accrued - deductions
    cost = round(accrued + base_social * 0.205)
    return {"accrued": accrued, "deductions": deductions, "net": net, "cost": cost}


def payroll_summary() -> str:
    wb = load_book()
    ws_e = wb["Empleados"]
    ws_n = wb["Novedades_Horas"]
    novelties = {
        str(row[1]): row
        for row in ws_n.iter_rows(min_row=2, values_only=True)
        if row[1]
    }
    active = 0
    totals = {"accrued": 0, "deductions": 0, "net": 0, "cost": 0}
    alerts = []
    for row in ws_e.iter_rows(min_row=2, values_only=True):
        emp_id, name, _, _, salary, *_rest = row
        is_active = row[11] == "Si"
        if not emp_id or not is_active:
            continue
        active += 1
        novelty = novelties.get(str(emp_id), [None, emp_id, HOURS_MONTH, 0, 0, 0, 0, 0, 0, 0])
        result = calculate_row(
            float(salary or 0),
            float(novelty[2] or 0),
            float(novelty[3] or 0),
            float(novelty[4] or 0),
            float(novelty[5] or 0),
            float(novelty[6] or 0),
            float(novelty[8] or 0),
            float(novelty[9] or 0),
        )
        for key in totals:
            totals[key] += result[key]
        if salary and salary < MIN_WAGE:
            alerts.append(f"{emp_id} salario bajo")
        if result["net"] < 0:
            alerts.append(f"{emp_id} neto negativo")
    return (
        "Resumen de nomina\n"
        f"Empleados activos: {active}\n"
        f"Total devengado: ${totals['accrued']:,.0f}\n"
        f"Total deducciones: ${totals['deductions']:,.0f}\n"
        f"Neto a pagar: ${totals['net']:,.0f}\n"
        f"Costo empresa estimado: ${totals['cost']:,.0f}\n"
        f"Alertas: {len(alerts)}"
    )


def simulate_raise(percent: float) -> str:
    base = payroll_summary()
    net_match = re.search(r"Neto a pagar: \$([0-9,]+)", base)
    current_net = int(net_match.group(1).replace(",", "")) if net_match else 0
    simulated = round(current_net * (1 + percent / 100))
    impact = simulated - current_net
    return (
        f"Simulacion aumento {percent:.2f}%\n"
        f"Neto actual aprox.: ${current_net:,.0f}\n"
        f"Neto simulado aprox.: ${simulated:,.0f}\n"
        f"Impacto aprox.: ${impact:,.0f}"
    )


def validate_workbook() -> str:
    wb = load_book()
    ws_e = wb["Empleados"]
    errors = []
    for row in range(2, ws_e.max_row + 1):
        emp_id = ws_e.cell(row, 1).value
        name = ws_e.cell(row, 2).value
        salary = ws_e.cell(row, 5).value
        if emp_id and not name:
            errors.append(f"{emp_id}: falta nombre")
        if emp_id and (salary is None or salary < MIN_WAGE):
            errors.append(f"{emp_id}: salario menor al minimo")
    if not errors:
        return "Validacion terminada: no encontre errores criticos en empleados."
    return "Validacion terminada:\n" + "\n".join(errors[:30])


def help_text() -> str:
    return (
        "Agente de nomina listo.\n\n"
        "Puedes escribirme:\n"
        "- enviame el excel\n"
        "- resumen de nomina\n"
        "- agrega 5 empleados con salario minimo\n"
        "- cambia salario EMP-0003 a 2500000\n"
        "- simula aumento del 10%\n"
        "- valida errores\n\n"
        "Cada cambio guarda copia en la carpeta backups."
    )


def handle_text(text: str) -> tuple[str, bool]:
    normalized = text.lower().strip()
    if normalized in {"/start", "ayuda", "help", "/help"}:
        return help_text(), False
    if "envi" in normalized and "excel" in normalized:
        return "Te envio el Excel actualizado.", True
    if "resumen" in normalized or "calcula nomina" in normalized or "calcular nomina" in normalized:
        return payroll_summary(), False
    if "valid" in normalized or "errores" in normalized:
        return validate_workbook(), False
    match = re.search(r"agreg\w*\s+(\d+)\s+emplead", normalized)
    if match and ("minimo" in normalized or "mínimo" in normalized):
        return add_minimum_wage_employees(int(match.group(1))), False
    match = re.search(r"salario\s+(emp-\d+)\s+(?:a|en)?\s*\$?\s*([\d.,]+)", normalized, re.IGNORECASE)
    if "cambia" in normalized and match:
        salary = int(re.sub(r"\D", "", match.group(2)))
        return change_salary(match.group(1).upper(), salary), False
    match = re.search(r"aumento\s+(?:del\s+)?(\d+(?:[.,]\d+)?)\s*%", normalized)
    if "simula" in normalized and match:
        return simulate_raise(float(match.group(1).replace(",", "."))), False
    return "No entendi la instruccion. Escribe 'ayuda' para ver ejemplos.", False


def run_bot() -> None:
    if not TOKEN:
        raise SystemExit("Falta TELEGRAM_BOT_TOKEN. Configuralo antes de ejecutar el agente.")
    offset = None
    print("Agente de Telegram para nomina iniciado.")
    while True:
        try:
            payload = {"timeout": 50}
            if offset:
                payload["offset"] = offset
            updates = request_json("getUpdates", payload).get("result", [])
            for update in updates:
                offset = update["update_id"] + 1
                message = update.get("message") or update.get("edited_message") or {}
                chat = message.get("chat", {})
                chat_id = chat.get("id")
                text = message.get("text", "")
                if not chat_id or not text:
                    continue
                print(f"Mensaje recibido de chat_id={chat_id}: {text}")
                if not authorize(chat_id):
                    send_message(chat_id, "No autorizado para operar este agente.")
                    continue
                try:
                    reply, attach_excel = handle_text(text)
                    send_message(chat_id, reply)
                    if attach_excel:
                        send_document(chat_id, EXCEL_PATH, "Nomina actualizada")
                except Exception as exc:
                    send_message(chat_id, f"Tuve un error procesando la instruccion: {exc}")
        except (urllib.error.URLError, TimeoutError) as exc:
            print(f"Conexion Telegram: {exc}. Reintentando...")
            time.sleep(5)


if __name__ == "__main__":
    run_bot()
